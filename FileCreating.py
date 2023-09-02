import easygui
import threading

from fpdf import FPDF
import pandas as excel
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as ImageEx
import resourses
from PIL import Image, ImageDraw, ImageFont


numberDetails = []
nameDetails = []
numDetails = []
imgDetails = []
OutDiamDetails = []
InnDiamDetails = []
LengthDetails = []
CostDetails = []
NData = []
Data = []
numDet = []
nameTable = []

def cleararray():
    numberDetails.clear()
    nameDetails.clear()
    numDetails.clear()
    imgDetails.clear()
    OutDiamDetails.clear()
    InnDiamDetails.clear()
    LengthDetails.clear()
    CostDetails.clear()
    NData.clear()
    Data.clear()
    numDet.clear()
    nameTable.clear()

def sizeimg(img):
    with Image.open(img) as img_1:
        img_1.load()
    return (img_1.height * 50) // img_1.width

def createPNG():
    img_monastery = Image.new('RGB', (168, 3076), color=('#FFF'))
    d = imgDetails[0:len(imgDetails) - 4]
    h = 10
    for i in range(4):
        filename_1 = d[i]
        with Image.open(filename_1) as img_1:
            img_1.load()
        h1 = (img_1.height * 150) // img_1.width
        img_monastery.paste(
            img_1.resize((150, h1)),
            (10, h),
        )
        font = ImageFont.truetype("Fonts/Circe-Regular.ttf", 40)
        draw_text = ImageDraw.Draw(img_monastery)
        draw_text.text((5, h + 10),
                       str(len(imgDetails) - i),
                       fill=('#000'), font=font)
        h = h + h1
    # показ изображения и его сохранение
    # img_monastery.show()
    img_monastery.save("ecx1.png")

    d = imgDetails[4:len(imgDetails)]
    h = 10
    hmass = []
    img = Image.new('RGB', (168, 5000), color=('#FFF'))
    for i in range(len(imgDetails) - 4):
        hmass.append(h)
        filename_1 = d[i]
        with Image.open(filename_1) as img_1:
            img_1.load()

        h1 = (img_1.height * 150) // img_1.width
        img.paste(
            img_1.resize((150, h1)),
            (10, h),
        )

        h = h + h1

    img = img.rotate(90, expand=True)
    for i in range(len(imgDetails) - 4):
        font = ImageFont.truetype("Fonts/Circe-Regular.ttf", 40)
        draw_text = ImageDraw.Draw(img)
        draw_text.text((hmass[i], 0),
                       str(len(imgDetails) - i - 4),
                       fill=('#000'), font=font)

    img.save("ecx2.png")
    img.close()


def createPDF(self, Path):
    from fpdf import FPDF
    pdf = FPDF(orientation="landscape")
    # директория где лежат системные шрифты OS Linux
    font_dir = 'Fonts'
    # добавляем TTF-шрифт, поддерживающий кириллицу.
    pdf.add_font("Serif", style="", fname=f"{font_dir}/ofont.ru_Times New Roman.ttf", uni=True)
    pdf.add_page()
    pdf.set_font("Serif", size=10)
    # высота ячейки
    line_height = pdf.font_size * 2.5

    pdf.set_x(50)
    print(nameTable[0])
    pdf.cell(237.5, line_height, nameTable[0], 1)
    pdf.ln(line_height)

    col_width = 158.3 / 4

    jojoreference = True
    for row in NData:
        pdf.set_x(50)
        for datum in row:
            pdf.cell(col_width, line_height, datum, 1)
        if jojoreference == 1:
            pdf.cell(79.2, line_height * 4, "Заметки:", 1)
            jojoreference = False
        pdf.ln(line_height)

    # одинаковая ширина ячеек
    col_width = 150 / 6
    for row in Data:
        pdf.set_x(50)
        pdf.cell(col_width / 3, line_height, row[0], 1)
        pdf.cell(col_width * 6, line_height, row[1], 1)
        pdf.cell(col_width * 1.5, line_height, row[2], 1)
        pdf.cell(col_width / 2, line_height, row[3], 1)
        pdf.cell(col_width / 2, line_height, row[4], 1)
        pdf.cell(col_width / 1.5, line_height, row[5], 1)
        pdf.ln(line_height)

    # Вывод картинок
    w = (168 * 200) // 3076
    pdf.image("ecx1.png", x=10, y=5, w=w, h=200)
    pdf.image("ecx2.png", x=20, y=190, w=300, h=10)
    # pdf.image("icon/RimeraLogo.png", x=210, y=170, w=50, h=10)
    pdf.output(Path)
def createExcel(self, Path):
    projectInfo = excel.DataFrame({
                        'NData': NData,
                        'Data': Data
                                  })
    infoCompon = excel.DataFrame({
        '№': numberDetails,
        'Название элемента': nameDetails,
        'Номер элемента': numDetails,
        'Наруж диам, мм': OutDiamDetails,
        'Внутр диам, мм': InnDiamDetails,
        'Длина, мм': LengthDetails
    })
    projectInfo.to_excel(Path)
    excel.set_option('max_colwidth', 120)
    excel.set_option('display.width', 500)
    with excel.ExcelWriter(Path) as writer:
        projectInfo.to_excel(writer, sheet_name='Sheet1', header=False, index=False, startcol=6, startrow=1)
        infoCompon.to_excel(writer, sheet_name='Sheet1', index=False, startcol=6, startrow=12)
    wb = load_workbook(Path)
    ws = wb.active
    r = 1
    for i in range(len(nameDetails)):
        img = str(imgDetails[i])
        logo = ImageEx(img)
        logo.width = 50
        logo.height = sizeimg(img)
        ws.add_image(logo, "B" + str(r))
        ws["A" + str(r)] = numberDetails[i]
        ws["C" + str(r)] = numDetails[i]
        r += logo.height // 15
    wb.save(Path)





def NewFile(self, FileType):
    print('f')
    match (FileType):
        case "PDF":
            createPNG()
            input_file = easygui.filesavebox("PDF-документ", "Сохранить как", "PDFSheet.pdf", filetypes=["*.pdf"])
            print(len(Data))
            if (input_file != None):
                """thread = threading.Thread(
                target=lambda: createPDF(self, input_file))
                thread.start()"""
                createPDF(self, input_file)
                print(777)
            cleararray()
        case "XLSX":
            input_file = easygui.filesavebox("Таблица Excel", "Сохранить как", "ExcelSheet.xlsx", filetypes=["*.xlsx"])
            if (input_file != None):
                createExcel(self, input_file)
            cleararray()
